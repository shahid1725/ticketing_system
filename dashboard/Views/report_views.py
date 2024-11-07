from django.shortcuts import render, redirect, get_object_or_404
from dashboard.models import *
from openpyxl import Workbook
import openpyxl
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

@login_required(login_url='admin_loginview')
def leads_report(request):
    start_date = None
    end_date = None
    leads_data = None

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        try:
            start_date_obj = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse("Invalid date format. Please use YYYY-MM-DD format.")

        leads_data = Leads.objects.filter(lead_date__range=[start_date_obj, end_date_obj]).order_by('lead_date')

        if 'export_excel' in request.POST:
            workbook = Workbook()
            worksheet = workbook.active

            headers = ['#', 'Full Name', 'Email', 'Mobile Number', 'Lead Date','Tag', 'Attended By']
            worksheet.append(headers)

            for idx, lead in enumerate(leads_data, start=1):
                formatted_lead_date = lead.lead_date.strftime('%d-%m-%Y') if lead.lead_date else ''
                row = [
                    idx, lead.full_name, lead.email, lead.mobile, formatted_lead_date,lead.tag.name if lead.tag else ''
                ]
                worksheet.append(row)

            column_widths = [5, 25, 15, 25, 15, 15,30]
            for i, width in enumerate(column_widths, start=1):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=leads_report.xlsx'
            workbook.save(response)
            return response

    return render(request, 'Report/LeadsReport.html', {'leads_data': leads_data, 'start_date': start_date, 'end_date': end_date})

@login_required(login_url='admin_loginview')
def sales_report(request):
    employees = CustomUser.objects.exclude(username='admin')
    resorts = Resorts.objects.all()
    
    selected_employee = request.POST.get('selected_employee') if request.method == 'POST' else request.GET.get('selected_employee')
    selected_resort = request.POST.get('selected_resort') if request.method == 'POST' else request.GET.get('selected_resort')
    start_date = request.POST.get('start_date') if request.method == 'POST' else request.GET.get('start_date')
    end_date = request.POST.get('end_date') if request.method == 'POST' else request.GET.get('end_date')
    employee_views_report = (request.POST.get('employee_views_report') == 'on') if request.method == 'POST' else (request.GET.get('employee_views_report') == 'on')

    filters = {}
    if start_date and end_date:
        filters['invoice_date__range'] = [start_date, end_date]
    if employee_views_report and selected_employee:
        filters['sales_person__username'] = selected_employee
    if selected_resort:
        filters['resort_name__id'] = selected_resort

    data = Invoice.objects.filter(**filters).order_by('checkin_date')
    
    paginator = Paginator(data, 200)
    page_number = request.GET.get('page')
    try:
        data = paginator.page(page_number)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)

    if 'export_excel' in request.POST:
        workbook = Workbook()
        worksheet = workbook.active

        headers = ['No', 'Date', 'Invoice No', 'Customer', 'Checkin Date', 'Price Quoted', 'Advance Amount', 'Due Date', 'Account ', 'Resort Name']
        worksheet.append(headers)

        column_widths = [5, 15, 15, 30, 15, 15, 15, 30, 15]
        for i, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        for idx, d in enumerate(data, start=1):
            account_name = d.account_name.account_name if d.account_name else ""
            resort_name = d.resort_name.resort_name if d.resort_name else ""
            row = [
                idx, d.invoice_date, d.invoice_number, 
                f"{d.customer.full_name}", d.checkin_date, d.package_price,
                d.recieved_price, d.due_date, 
                account_name, resort_name
            ]
            worksheet.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
        workbook.save(response)
        return response

    return render(request, 'Report/Report.html', {
        'data': data,
        'employees': employees,
        'selected_employee': selected_employee,
        'resorts': resorts,
        'selected_resort': selected_resort,
        'start_date': start_date,
        'end_date': end_date,
        'employee_views_report': employee_views_report,
    })


@login_required(login_url='admin_loginview')
def voucher_report(request):
    employees = CustomUser.objects.exclude(username='admin')
    resorts = Resorts.objects.all()

    selected_employee = request.POST.get('selected_employee') if request.method == 'POST' else request.GET.get(
        'selected_employee')
    selected_resort = request.POST.get('selected_resort') if request.method == 'POST' else request.GET.get(
        'selected_resort')
    start_date = request.POST.get('start_date') if request.method == 'POST' else request.GET.get('start_date')
    end_date = request.POST.get('end_date') if request.method == 'POST' else request.GET.get('end_date')
    employee_views_report = (request.POST.get('employee_views_report') == 'on') if request.method == 'POST' else (
                request.GET.get('employee_views_report') == 'on')

    filters = {}
    if start_date and end_date:
        filters['voucher_date__range'] = [start_date, end_date]
    if employee_views_report and selected_employee:
        filters['sales_person__username'] = selected_employee
    if selected_resort:
        filters['resort_name__id'] = selected_resort

    data = Voucher.objects.filter(**filters).order_by('checkin_date')

    paginator = Paginator(data, 200)
    page_number = request.GET.get('page')
    try:
        data = paginator.page(page_number)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)

    if 'export_excel' in request.POST:
        workbook = Workbook()
        worksheet = workbook.active

        headers = ['No', 'Date', 'Voucher No', 'Customer', 'Checkin Date', 'Price Quoted', 'Advance Amount', 'Travel', 'Account Number',
                   'Resort Name']
        worksheet.append(headers)

        column_widths = [5, 15, 15, 30, 15, 15, 15, 30, 15]
        for i, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        for idx, d in enumerate(data, start=1):
            account_name = d.account_name.account_name if d.account_name else ""
            resort_name = d.resort_name.resort_name if d.resort_name else ""
            row = [
                idx, d.voucher_date, d.voucher_number,
                f"{d.customer.full_name}", d.checkin_date, d.package_price,
                d.recieved_price, d.travel,
                account_name, resort_name
            ]
            worksheet.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=voucher_report.xlsx'
        workbook.save(response)
        return response

    return render(request, 'Report/VoucherReport.html', {
        'data': data,
        'employees': employees,
        'selected_employee': selected_employee,
        'resorts': resorts,
        'selected_resort': selected_resort,
        'start_date': start_date,
        'end_date': end_date,
        'employee_views_report': employee_views_report,
    })

def profit_report(request):
    employees = CustomUser.objects.exclude(username='admin')
    resorts = Resorts.objects.all()

    selected_employee = None
    selected_resort = None
    start_date = None
    end_date = None
    employee_views_report = False  # Default value for the checkbox

    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        selected_employee = request.POST.get('selected_employee')
        selected_resort = request.POST.get('selected_resort')
        employee_views_report = request.POST.get('employee_views_report')
    else:
        # If it's a GET request, check if the checkbox is initially checked
        employee_views_report = request.GET.get('employee_views_report')

    # Initialize empty lists for voucher and invoice data
    voucher_data = []
    invoice_data = []

    # Fetch vouchers based on filters
    voucher_filters = {'checkin_date__range': [start_date, end_date]}
    if employee_views_report:
        voucher_filters['sales_person__username'] = selected_employee
    if selected_resort:
        voucher_filters['resort_name__id'] = selected_resort

    voucher_data = Voucher.objects.filter(**voucher_filters).order_by('checkin_date')

    # Fetch invoices based on filters
    invoice_filters = {'checkin_date__range': [start_date, end_date]}
    if employee_views_report:
        invoice_filters['sales_person__username'] = selected_employee
    if selected_resort:
        invoice_filters['resort_name__id'] = selected_resort

    invoice_data = Invoice.objects.filter(**invoice_filters).order_by('checkin_date')

    # Pagination for voucher data
    voucher_paginator = Paginator(voucher_data, 500)  # 10 vouchers per page
    voucher_page_number = request.GET.get('voucher_page')
    try:
        voucher_data = voucher_paginator.page(voucher_page_number)
    except PageNotAnInteger:
        voucher_data = voucher_paginator.page(1)
    except EmptyPage:
        voucher_data = voucher_paginator.page(voucher_paginator.num_pages)

    # Pagination for invoice data
    invoice_paginator = Paginator(invoice_data, 500)  # 10 invoices per page
    invoice_page_number = request.GET.get('invoice_page')
    try:
        invoice_data = invoice_paginator.page(invoice_page_number)
    except PageNotAnInteger:
        invoice_data = invoice_paginator.page(1)
    except EmptyPage:
        invoice_data = invoice_paginator.page(invoice_paginator.num_pages)

    # Handle Excel export
    if 'export_excel' in request.POST:
        workbook = Workbook()
        worksheet_vouchers = workbook.create_sheet(title='Vouchers')
        worksheet_invoices = workbook.create_sheet(title='Invoices')

        # Headers for vouchers
        voucher_headers = [
            '#', 'Date', 'Voucher No', 'Customer', 'Mobile Number', 'Resort Name', 'Profit',
            'Price Quoted', 'Advance Amount', 'Balance Amount', 'Sales Person', 'Account Details', 'No of nights', 'No of rooms'
        ]
        worksheet_vouchers.append(voucher_headers)

        column_widths = [5, 15, 15, 30, 15, 30, 20, 15, 20, 15, 15, 30]
        for i, width in enumerate(column_widths, start=1):
            worksheet_vouchers.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        total_voucher_profit = 0
        for idx, d in enumerate(voucher_data, start=1):
            recieved_price = float(d.recieved_price) if isinstance(d.recieved_price, str) else d.recieved_price
            pending_price = float(d.pending_price) if isinstance(d.pending_price, str) else d.pending_price
            account_name = str(d.account_name) if d.account_name else ""
            profit = float(d.profit) if d.profit else 0
            total_voucher_profit += profit
            resort_name = str(d.resort_name) if d.resort_name else ""
            row = [
                idx, d.checkin_date, d.voucher_number,
                f"{d.customer.salutation}. {d.customer.full_name}", d.customer.contact_number, resort_name, profit,
                f"{d.total_amount:.0f}", f"{recieved_price:.0f}",
                f"{pending_price:.0f}", d.sales_person.username,
                account_name, f"{d.number_of_nights}", f"{d.number_of_rooms}"
            ]
            worksheet_vouchers.append(row)

        total_voucher_row = [''] * 5 + ['Total Profit', f"{total_voucher_profit:.0f}"]
        worksheet_vouchers.append(total_voucher_row)

        # Headers for invoices
        invoice_headers = [
            '#', 'Date', 'Invoice No', 'Customer', 'Mobile Number', 'Resort Name', 'Profit',
            'Received Price', 'Total Amount', 'Balance Amount', 'Sales Person', 'Account Details', 'No of nights',
            'No of rooms'
        ]
        worksheet_invoices.append(invoice_headers)

        column_widths = [5, 15, 15, 30, 15, 30, 20, 15, 20, 15, 15, 30]
        for i, width in enumerate(column_widths, start=1):
            worksheet_invoices.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        total_invoice_profit = 0
        for idx, d in enumerate(invoice_data, start=1):
            account_name = str(d.account_name) if d.account_name else ""
            resort_name = str(d.resort_name) if d.resort_name else ""
            profit = float(d.profit) if d.profit else 0
            total_invoice_profit += profit
            row = [
                idx, d.checkin_date, d.invoice_number,
                f"{d.customer.salutation}. {d.customer.full_name}", d.customer.contact_number,
                resort_name, f"{profit:.0f}", f"{d.recieved_price:.0f}", f"{d.total_amount:.0f}",
                f"{d.pending_price:.0f}", d.sales_person.username, account_name,
                f"{d.number_of_nights}", f"{d.number_of_rooms}"
            ]
            worksheet_invoices.append(row)

        total_invoice_row = [''] * 5 + ['Total Profit', f"{total_invoice_profit:.0f}"]
        worksheet_invoices.append(total_invoice_row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=profit_report.xlsx'
        workbook.save(response)
        return response

    return render(request, 'Report/Profit_report.html', {
        'voucher_data': voucher_data,
        'voucher_page_number': voucher_page_number,
        'voucher_paginator': voucher_paginator,
        'invoice_data': invoice_data,
        'invoice_page_number': invoice_page_number,
        'invoice_paginator': invoice_paginator,
        'employees': employees,
        'selected_employee': selected_employee,
        'resorts': resorts,
        'selected_resort': selected_resort,
        'start_date': start_date,
        'end_date': end_date,
        'employee_views_report': employee_views_report,
    })